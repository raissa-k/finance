"""Spreadsheet/CSV import that creates Obligations from a mappable column format.

Unlike the transaction ledger import, this is create-only + dedup-block (see
app/obligation_dedup.py) rather than diff-and-reconcile: an obligation either
gets created, or created-but-blocked as a likely duplicate for the user to
review. Parsing is pure; classify/apply need a DB session.
"""

from __future__ import annotations

import csv
import io
import re
from datetime import date, datetime, timezone
from typing import Optional

from sqlalchemy import func
from sqlalchemy.orm import Session

from app.models import Category, Obligation, ObligationGroup, ObligationOccurrence, Payee
from app.obligation_dedup import detect_duplicate_obligation, detect_duplicate_occurrence
from app.obligation_helpers import derive_period

# Target fields a column can be mapped to.
TARGET_FIELDS = [
    "name",
    "amount",
    "due_date",
    "period",
    "category",
    "payee",
    "recurrence",
    "is_recurring",
    "paid",
    "direction",
    "note",
]

_TRUTHY = {"true", "1", "yes", "sim", "x", "✓", "pago", "paid", "recebido", "received"}
_RECURRENCE_MAP = {
    "monthly": "monthly", "mensal": "monthly", "mes": "monthly", "mês": "monthly",
    "weekly": "weekly", "semanal": "weekly",
    "yearly": "yearly", "annual": "yearly", "anual": "yearly",
    "none": None, "nao": None, "não": None, "one-off": None, "one time": None,
    "unico": None, "único": None,
}
# Substring match (not exact) since real-world values are free text like
# "Despesa Fixa" / "Despesa Variável" / "Receita".
_RECEIVABLE_HINTS = ("receita", "income", "revenue", "receivable", "recebi")
_PAYABLE_HINTS = ("despesa", "expense", "payable", "bill")


def _resolve_direction(value) -> Optional[str]:
    if value is None or str(value).strip() == "":
        return None
    text = str(value).strip().lower()
    if any(hint in text for hint in _RECEIVABLE_HINTS):
        return "receivable"
    if any(hint in text for hint in _PAYABLE_HINTS):
        return "payable"
    return None


_MONTH_NAME_TO_NUM = {
    "jan": 1, "fev": 2, "feb": 2, "mar": 3, "abr": 4, "apr": 4, "mai": 5, "may": 5,
    "jun": 6, "jul": 7, "ago": 8, "aug": 8, "set": 9, "sep": 9, "out": 10, "oct": 10,
    "nov": 11, "dez": 12, "dec": 12,
}


def _parse_period(value) -> Optional[str]:
    """Lenient "YYYY-MM" parse for a free-text month/year column (e.g. a
    spreadsheet's "Mês" column: "2026 07", "07/2026", "Jul 2026", ...)."""
    if value is None or str(value).strip() == "":
        return None
    if isinstance(value, (datetime, date)):
        return value.strftime("%Y-%m")
    text = str(value).strip()

    m = re.match(r"^(\d{4})[-/\s](\d{1,2})$", text)
    if m and 1 <= int(m.group(2)) <= 12:
        return f"{int(m.group(1)):04d}-{int(m.group(2)):02d}"

    m = re.match(r"^(\d{1,2})[-/\s](\d{4})$", text)
    if m and 1 <= int(m.group(1)) <= 12:
        return f"{int(m.group(2)):04d}-{int(m.group(1)):02d}"

    m = re.match(r"^([A-Za-zçÇãÃéÉ]+)\s+(\d{4})$", text)
    if m:
        month_num = _MONTH_NAME_TO_NUM.get(m.group(1).strip().lower()[:3])
        if month_num:
            return f"{int(m.group(2)):04d}-{month_num:02d}"

    m = re.match(r"^(\d{4})\s+([A-Za-zçÇãÃéÉ]+)$", text)
    if m:
        month_num = _MONTH_NAME_TO_NUM.get(m.group(2).strip().lower()[:3])
        if month_num:
            return f"{int(m.group(1)):04d}-{month_num:02d}"

    return None


# ── Parsing (generic, mirrors the reconciling importer's file-reading layer) ──

def _sheet_rows(content: bytes, sheet_name: Optional[str], header_row: int):
    """Yield (headers, list-of-row-dicts) for an xlsx file using openpyxl."""
    import openpyxl

    wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True, read_only=True)
    ws = wb[sheet_name] if sheet_name and sheet_name in wb.sheetnames else wb[wb.sheetnames[0]]

    all_rows = list(ws.iter_rows(values_only=True))
    if header_row < 1 or header_row > len(all_rows):
        return [], []
    headers = [str(h).strip() if h is not None else "" for h in all_rows[header_row - 1]]
    data = []
    for raw in all_rows[header_row:]:
        row = {headers[i]: raw[i] for i in range(len(headers)) if i < len(raw)}
        data.append(row)
    return headers, data


def parse_file(content: bytes, file_type: str, sheet_name: Optional[str], header_row: int):
    """Return (headers, rows) where each row is {header: cell_value}."""
    if (file_type or "").lower() == "csv":
        text = content.decode("utf-8-sig")
        reader = csv.reader(io.StringIO(text))
        all_rows = list(reader)
        if header_row < 1 or header_row > len(all_rows):
            return [], []
        headers = [str(h).strip() for h in all_rows[header_row - 1]]
        rows = []
        for raw in all_rows[header_row:]:
            rows.append({headers[i]: (raw[i] if i < len(raw) else None) for i in range(len(headers))})
        return headers, rows
    return _sheet_rows(content, sheet_name, header_row)


def _header_cell_count(ws, header_row: int) -> int:
    rows = list(ws.iter_rows(min_row=header_row, max_row=header_row, values_only=True))
    if not rows:
        return 0
    return sum(1 for c in rows[0] if c is not None and str(c).strip() != "")


def _best_sheet(wb, sheets, header_row: int) -> str:
    """Pick the sheet whose header row has the most non-empty cells."""
    best, best_count = sheets[0], -1
    for name in sheets:
        count = _header_cell_count(wb[name], header_row)
        if count > best_count:
            best, best_count = name, count
    return best


def analyze(content: bytes, file_type: str, sheet_name: Optional[str] = None, header_row: int = 1):
    """Inspect an upload: available sheets, headers and a few sample rows."""
    header_row = header_row or 1
    if (file_type or "").lower() == "csv":
        headers, rows = parse_file(content, "csv", None, header_row)
        return {"sheets": [], "sheet_name": None, "headers": headers, "sample": rows[:5]}

    import openpyxl

    wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True, read_only=True)
    sheets = wb.sheetnames
    target = sheet_name if (sheet_name and sheet_name in sheets) else _best_sheet(wb, sheets, header_row)
    headers, rows = _sheet_rows(content, target, header_row)
    sample = [
        {k: (v.isoformat() if isinstance(v, (datetime, date)) else v) for k, v in r.items()}
        for r in rows[:5]
    ]
    return {"sheets": sheets, "sheet_name": target, "headers": headers, "sample": sample}


def _parse_date(value, date_format: Optional[str]) -> Optional[date]:
    if value is None or value == "":
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    text = str(value).strip()
    if date_format:
        try:
            return datetime.strptime(text, date_format).date()
        except ValueError:
            pass
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%m/%d/%Y", "%Y/%m/%d", "%d-%m-%Y", "%Y-%m-%dT%H:%M:%S"):
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            continue
    raise ValueError(f"Unrecognized date: {value!r}")


def _parse_amount(value, decimal_separator: str) -> float:
    if value is None or value == "":
        raise ValueError("Missing amount")
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).strip()
    for sym in ("R$", "US$", "$", "£", "€", " "):
        text = text.replace(sym, "")
    if decimal_separator == ",":
        text = text.replace(".", "").replace(",", ".")
    else:
        text = text.replace(",", "")
    if text in ("", "-", "+"):
        raise ValueError("Missing amount")
    return float(text)


def _clean_str(value) -> Optional[str]:
    if value is None:
        return None
    text = str(value).strip()
    return text or None


def _norm(text: Optional[str]) -> str:
    return " ".join((text or "").strip().lower().split())


def _is_paid(row, mapping) -> bool:
    val = row.get(mapping["paid"])
    if isinstance(val, bool):
        return val
    return str(val).strip().lower() in _TRUTHY


def _resolve_recurrence(row, mapping, default_recurrence: Optional[str]) -> tuple[bool, Optional[str]]:
    if "recurrence" in mapping:
        raw = row.get(mapping["recurrence"])
        if raw is not None and str(raw).strip() != "":
            recurrence = _RECURRENCE_MAP.get(str(raw).strip().lower(), default_recurrence)
            return bool(recurrence), recurrence

    if "is_recurring" in mapping:
        raw = row.get(mapping["is_recurring"])
        if raw is not None and str(raw).strip() != "":
            truthy = str(raw).strip().lower() in _TRUTHY
            return truthy, (default_recurrence if truthy else None)

    # Neither column mapped/filled -> the primary use case is bulk-importing
    # recurring bills, so default every row to the format's cadence.
    return bool(default_recurrence), default_recurrence


# ── Normalization ─────────────────────────────────────────────────────────────

def resolve_config(fmt: dict) -> dict:
    """Normalize a saved/inline format definition into a parsing config."""
    mapping = {}
    for f in fmt.get("fields", []):
        target = f.get("target_field")
        source = f.get("source_column")
        if target and source:
            mapping[target] = source
    return {
        "file_type": fmt.get("file_type", "xlsx"),
        "sheet_name": fmt.get("sheet_name"),
        "header_row": fmt.get("header_row", 1) or 1,
        "date_format": fmt.get("date_format"),
        "decimal_separator": fmt.get("decimal_separator", ".") or ".",
        "default_recurrence": fmt.get("default_recurrence") or None,
        "default_category_id": fmt.get("default_category_id"),
        "mapping": mapping,
    }


def build_records(rows, config):
    """Normalize raw rows into obligation records; returns (records, errors)."""
    mapping = config["mapping"]
    records, errors = [], []

    if "name" not in mapping or "amount" not in mapping:
        errors.append({"row": 0, "message": "The format must map at least 'name' and 'amount'."})
        return records, errors

    for idx, row in enumerate(rows):
        if all((row.get(col) is None or str(row.get(col)).strip() == "") for col in mapping.values()):
            continue
        try:
            name = _clean_str(row.get(mapping["name"]))
            if not name:
                continue  # a row with no name is treated as blank/ignored
            amount = _parse_amount(row.get(mapping["amount"]), config["decimal_separator"])

            due_date = None
            if "due_date" in mapping:
                due_date = _parse_date(row.get(mapping["due_date"]), config["date_format"])

            is_recurring, recurrence = _resolve_recurrence(row, mapping, config["default_recurrence"])

            records.append(
                {
                    "row": idx + 1,
                    "name": name,
                    "amount": abs(amount),
                    "due_date": due_date,
                    "period": _parse_period(row.get(mapping["period"])) if "period" in mapping else None,
                    "is_recurring": is_recurring,
                    "recurrence": recurrence,
                    "category": _clean_str(row.get(mapping["category"])) if "category" in mapping else None,
                    "payee": _clean_str(row.get(mapping["payee"])) if "payee" in mapping else None,
                    "note": _clean_str(row.get(mapping["note"])) if "note" in mapping else None,
                    "paid": _is_paid(row, mapping) if "paid" in mapping else False,
                    "direction": _resolve_direction(row.get(mapping["direction"])) if "direction" in mapping else None,
                }
            )
        except ValueError as e:
            errors.append({"row": idx + 1, "message": str(e)})
    return records, errors


# ── Classify (resolve category/payee text) ─────────────────────────────────────

def classify_and_stage(records: list[dict], db: Session) -> list[dict]:
    """Resolve each row's category/payee text to an existing id, if any."""
    results = []
    for rec in records:
        category_id = None
        if rec["category"]:
            cat = (
                db.query(Category)
                .filter(func.lower(func.trim(Category.name)) == _norm(rec["category"]))
                .first()
            )
            category_id = cat.category_id if cat else None

        payee_id = None
        if rec["payee"]:
            payee = (
                db.query(Payee)
                .filter(func.lower(func.trim(Payee.name)) == _norm(rec["payee"]))
                .first()
            )
            payee_id = payee.payee_id if payee else None

        results.append({"record": rec, "category_id": category_id, "payee_id": payee_id})
    return results


# ── Group (one Obligation per distinct bill, one Occurrence per row) ──────────

def group_records(classified: list[dict], db: Session) -> list[dict]:
    """Group rows sharing a normalized (name, category) into one obligation.

    A spreadsheet that lists one row per period for the same recurring bill
    (e.g. 12 monthly "Aluguel" rows) must become ONE Obligation with 12
    occurrences -- not 12 near-identical Obligations that immediately flag
    each other as duplicates.

    When the bill's (name, category) already matches an existing unblocked
    Obligation, this does NOT flag the whole group as one big "duplicate" --
    re-importing "Salário" a year later, with entirely different months, is
    the normal/expected way this feature gets used, not a mistake. Instead
    each occurrence is individually checked (detect_duplicate_occurrence)
    against that EXISTING obligation's own occurrences: only a row whose
    period genuinely overlaps something already there is a duplicate; the
    rest are attached to it at apply time (see attach_to_obligation_id).
    """
    order: list[tuple] = []
    groups: dict[tuple, dict] = {}
    for c in classified:
        rec = c["record"]
        key = (_norm(rec["name"]), c["category_id"])
        if key not in groups:
            groups[key] = {
                "name": rec["name"],
                "category_id": c["category_id"],
                "category_raw": rec["category"],
                "payee_id": c["payee_id"],
                "payee_raw": rec["payee"],
                "rows": [],
            }
            order.append(key)
        groups[key]["rows"].append(c)

    result = []
    for key in order:
        g = groups[key]
        rows = sorted(g["rows"], key=lambda c: (c["record"]["due_date"] is None, c["record"]["due_date"]))
        first_rec = rows[0]["record"]

        if len(rows) > 1:
            is_recurring = True
            recurrence = first_rec["recurrence"] or "monthly"
        else:
            is_recurring = first_rec["is_recurring"]
            recurrence = first_rec["recurrence"]

        # A bill's direction (payable/receivable) doesn't vary month to month,
        # so one row specifying it (however it's mapped in the sheet) settles
        # it for the whole group; default to "payable" like the model itself.
        direction = next((c["record"]["direction"] for c in rows if c["record"]["direction"]), None) or "payable"

        category_id = g["category_id"]

        # If this bill's name matches an EXISTING ObligationGroup exactly,
        # the group's settings win over whatever this row/import resolved to
        # -- see ObligationGroup's docstring for why groups exist at all.
        obligation_group_id = None
        matched_group = (
            db.query(ObligationGroup).filter(func.lower(func.trim(ObligationGroup.name)) == _norm(g["name"])).first()
        )
        if matched_group:
            obligation_group_id = matched_group.obligation_group_id
            category_id = matched_group.category_id
            direction = matched_group.direction
            if matched_group.recurrence:
                is_recurring = True
                recurrence = matched_group.recurrence

        final_recurrence = recurrence if is_recurring else None
        occurrences = [c["record"] for c in rows]

        dup = detect_duplicate_obligation(db, g["name"], category_id)
        for occ in occurrences:
            dup_occ = detect_duplicate_occurrence(db, dup.obligation_id, occ["due_date"], final_recurrence) if dup else None
            occ["is_duplicate"] = dup_occ is not None
            occ["duplicate_of_occurrence_id"] = dup_occ.obligation_occurrence_id if dup_occ else None

        result.append(
            {
                "name": g["name"],
                "category_id": category_id,
                "category_raw": g["category_raw"],
                "payee_id": g["payee_id"],
                "payee_raw": g["payee_raw"],
                "direction": direction,
                "obligation_group_id": obligation_group_id,
                "obligation_group_name": matched_group.name if matched_group else None,
                "is_recurring": is_recurring,
                "recurrence": final_recurrence,
                "occurrences": occurrences,
                "first_row": first_rec["row"],
                # Existing, unblocked Obligation this whole group matches by
                # (name, category) -- when set, apply() attaches non-conflicting
                # occurrences to it directly instead of creating a new one.
                "attach_to_obligation_id": dup.obligation_id if dup else None,
                "duplicate_of_obligation_id": dup.obligation_id if dup else None,
                "duplicate_reason": (
                    f"Matches existing obligation #{dup.obligation_id} ({dup.name}) -- new occurrences will "
                    "attach to it" if dup else None
                ),
            }
        )
    return result


def build_preview(groups: list[dict]) -> dict:
    rows = []
    new_obligations = 0
    attached_obligations = 0
    new_occurrences = 0
    duplicate_occurrences = 0
    for g in groups:
        if g["attach_to_obligation_id"] is not None:
            attached_obligations += 1
        else:
            new_obligations += 1
        for occ in g["occurrences"]:
            if occ["is_duplicate"]:
                duplicate_occurrences += 1
            else:
                new_occurrences += 1
            rows.append(
                {
                    "row": occ["row"],
                    "name": g["name"],
                    "amount": occ["amount"],
                    "due_date": occ["due_date"].isoformat() if occ["due_date"] else None,
                    "period": derive_period(occ["due_date"], occ["period"]),
                    "is_recurring": g["is_recurring"],
                    "recurrence": g["recurrence"],
                    "category_raw": g["category_raw"],
                    "category_id": g["category_id"],
                    "payee_raw": g["payee_raw"],
                    "payee_id": g["payee_id"],
                    "direction": g["direction"],
                    "obligation_group_id": g["obligation_group_id"],
                    "obligation_group_name": g["obligation_group_name"],
                    "note": occ["note"],
                    "paid": occ["paid"],
                    "paid_date": occ["due_date"].isoformat() if occ["paid"] and occ["due_date"] else None,
                    "occurrence_of_row": g["first_row"],
                    # Per-OCCURRENCE truth: does *this* row's period genuinely
                    # collide with something already in the DB? A bill name
                    # matching an existing obligation does NOT by itself make
                    # every row a duplicate -- most re-imports add new months.
                    "is_duplicate": occ["is_duplicate"],
                    "attach_to_obligation_id": g["attach_to_obligation_id"],
                    "duplicate_of_obligation_id": g["duplicate_of_obligation_id"],
                    "duplicate_reason": g["duplicate_reason"],
                }
            )

    unmatched_categories = sorted({g["category_raw"] for g in groups if g["category_raw"] and g["category_id"] is None})
    unmatched_payees = sorted({g["payee_raw"] for g in groups if g["payee_raw"] and g["payee_id"] is None})
    unmatched_group_names = sorted({g["name"] for g in groups if g["obligation_group_id"] is None})

    return {
        "summary": {
            "new_obligations": new_obligations,
            "attached_obligations": attached_obligations,
            "new_occurrences": new_occurrences,
            "duplicate_occurrences": duplicate_occurrences,
            "errors": 0,
            "rows": len(rows),
        },
        "rows": rows,
        "unmatched_categories": unmatched_categories,
        "unmatched_payees": unmatched_payees,
        "unmatched_group_names": unmatched_group_names,
    }


# ── Apply (create one Obligation + its occurrences per group) ─────────────────

def apply_records(
    groups: list[dict],
    resolutions: dict,
    config: dict,
    db: Session,
    format_id: Optional[int] = None,
    skip_duplicates: bool = False,
) -> dict:
    cache_cat: dict[str, int] = {}
    cache_payee: dict[str, int] = {}

    def ensure_category(name: Optional[str], parent_name: Optional[str] = None) -> Optional[int]:
        if not name:
            return None
        k = _norm(name)
        if k in cache_cat:
            return cache_cat[k]
        obj = db.query(Category).filter(func.lower(func.trim(Category.name)) == k).first()
        if not obj:
            parent_id = None
            if parent_name:
                parent_obj = (
                    db.query(Category).filter(func.lower(func.trim(Category.name)) == _norm(parent_name)).first()
                )
                parent_id = parent_obj.category_id if parent_obj else None
            obj = Category(name=name, parent_category_id=parent_id, is_hidden=False)
            db.add(obj)
            db.flush()
        cache_cat[k] = obj.category_id
        return obj.category_id

    def ensure_payee(name: Optional[str]) -> Optional[int]:
        if not name:
            return None
        k = _norm(name)
        if k in cache_payee:
            return cache_payee[k]
        obj = db.query(Payee).filter(func.lower(func.trim(Payee.name)) == k).first()
        if not obj:
            obj = Payee(name=name)
            db.add(obj)
            db.flush()
        cache_payee[k] = obj.payee_id
        return obj.payee_id

    created = attached = 0
    occurrences_created = occurrences_blocked = 0
    obligation_ids: list[int] = []

    def add_occurrence(obligation_id: int, occ: dict) -> None:
        nonlocal occurrences_created, occurrences_blocked
        occurrence = ObligationOccurrence(
            obligation_id=obligation_id,
            due_date=occ["due_date"],
            period=derive_period(occ["due_date"], occ["period"]),
            estimated_amount=occ["amount"],
            paid=occ["paid"],
            paid_at=datetime.now(timezone.utc) if occ["paid"] else None,
            # Backfilled historical data -- "today" (the import date) would be
            # meaningless as a paid/received date, so default to the
            # occurrence's own due date instead; blank when unpaid, and
            # editable by hand afterward either way.
            paid_date=occ["due_date"] if occ["paid"] else None,
            note=occ["note"],
            source="import",
            created_via_format_id=format_id,
        )
        dup_occ = detect_duplicate_occurrence(db, obligation_id, occ["due_date"], g["recurrence"])
        if dup_occ:
            occurrence.is_blocked = True
            occurrence.duplicate_of_occurrence_id = dup_occ.obligation_occurrence_id
            occurrence.blocked_reason = (
                f"Duplicate of occurrence #{dup_occ.obligation_occurrence_id} ({dup_occ.due_date})"
            )
            occurrences_blocked += 1
        else:
            occurrences_created += 1
        db.add(occurrence)
        db.flush()

    for g in groups:
        override = resolutions.get(str(g["first_row"]), {}) if resolutions else {}

        category_id = override.get("category_id", g["category_id"])
        if category_id is None and override.get("category_name"):
            # AI proposed a new (translated) category name during "Preview with
            # AI" -- create/reuse that instead of falling back to the raw,
            # untranslated spreadsheet text below.
            category_id = ensure_category(override["category_name"], override.get("category_parent"))
        if category_id is None and g["category_raw"]:
            category_id = ensure_category(g["category_raw"])
        if category_id is None:
            category_id = config.get("default_category_id")

        payee_id = override.get("payee_id", g["payee_id"])
        if payee_id is None and g["payee_raw"]:
            payee_id = ensure_payee(g["payee_raw"])

        direction = g["direction"]
        is_recurring = g["is_recurring"]
        recurrence = g["recurrence"]
        obligation_group_id = g["obligation_group_id"]

        # An AI-suggested group match chosen in the preview step (for names
        # that didn't exact-match one already applied in group_records)
        # applies the SAME rule: the group's settings win over the row's own.
        resolved_group_id = override.get("obligation_group_id")
        if resolved_group_id:
            matched_group = db.get(ObligationGroup, resolved_group_id)
            if matched_group:
                obligation_group_id = matched_group.obligation_group_id
                category_id = matched_group.category_id
                direction = matched_group.direction
                if matched_group.recurrence:
                    is_recurring = True
                    recurrence = matched_group.recurrence

        # Re-check now that category may have just been created/overridden --
        # the group() pass ran this against whatever existed at preview time.
        dup = detect_duplicate_obligation(db, g["name"], category_id)
        if dup and skip_duplicates:
            continue

        if dup is not None:
            # Re-importing the same recurring bill (same name+category) months
            # or years later is the normal way this feature is used -- attach
            # the new, non-conflicting occurrences straight onto the EXISTING
            # obligation instead of spinning up another one to reconcile by
            # hand. Only a row whose own period actually collides gets
            # blocked (below, same per-occurrence check as the new-obligation
            # path) -- never the whole group.
            ob_id = dup.obligation_id
            attached += 1
        else:
            ob = Obligation(
                name=g["name"],
                category_id=category_id,
                payee_id=payee_id,
                obligation_group_id=obligation_group_id,
                is_recurring=is_recurring,
                recurrence=recurrence,
                estimated_amount=g["occurrences"][0]["amount"],
                direction=direction,
                note=g["occurrences"][0]["note"],
                source="import",
                created_via_format_id=format_id,
            )
            db.add(ob)
            db.flush()
            ob_id = ob.obligation_id
            created += 1

        for occ in g["occurrences"]:
            add_occurrence(ob_id, occ)

        obligation_ids.append(ob_id)

    db.commit()
    return {
        "ok": True,
        "created": created,
        "attached": attached,
        "occurrences_created": occurrences_created,
        "occurrences_blocked": occurrences_blocked,
        "obligation_ids": obligation_ids,
        "errors": [],
    }
